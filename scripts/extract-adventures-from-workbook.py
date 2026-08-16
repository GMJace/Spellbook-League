import csv
import json
import re
import sys
from collections import OrderedDict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


HELPER_SHEETS = {
    "Table of Contents",
    "READ ME",
    "IndexHelper",
}

ITEM_SHEETS = {"Unique Items"}

ADVENTURE_HEADER_KEYWORDS = [
    "adventure code",
    "module code",
    "adventure title",
    "level range",
    "runtime",
    "consumables",
    "magic item",
    "spellbooks",
    "story award",
    "treasure",
    "page numbers",
    "chapter",
]

RARITY_BUCKETS = {
    "common": "commonMagicItems",
    "uncommon": "uncommonMagicItems",
    "rare": "rareMagicItems",
    "very rare": "veryRareMagicItems",
    "legendary": "legendaryMagicItems",
}

CODE_PATTERNS = [
    re.compile(r"^adventure$"),
    re.compile(r"^chapter$"),
    re.compile(r"adventure code"),
    re.compile(r"module code"),
]
TITLE_PATTERNS = [
    re.compile(r"adventure title"),
    re.compile(r"chapter section title"),
    re.compile(r"chapter title"),
    re.compile(r"^title$"),
]
LEVEL_PATTERNS = [re.compile(r"level range"), re.compile(r"^tier$"), re.compile(r"^level$")]
RECOMMENDED_PATTERNS = [re.compile(r"recom"), re.compile(r"optimized for")]
APL_PATTERNS = [re.compile(r"^apl$")]
RUNTIME_PATTERNS = [re.compile(r"runtime")]
CONSUMABLE_PATTERNS = [re.compile(r"^consumables")]
MAGIC_ITEM_PATTERNS = [re.compile(r"^magic items?$"), re.compile(r"^magic item ")]
SPELLBOOK_PATTERNS = [re.compile(r"spellbooks"), re.compile(r"scrolls")]
STORY_AWARD_PATTERNS = [re.compile(r"story award"), re.compile(r"^specials$")]
PAGE_NUMBER_PATTERNS = [re.compile(r"page numbers")]
FIXED_GOLD_PATTERNS = [re.compile(r"fixed treasure"), re.compile(r"^treasure gp$")]
ROLLED_GOLD_PATTERNS = [re.compile(r"rolled treasure")]
LEGACY_GOLD_PATTERNS = [re.compile(r"treasure gp legacy")]
AUTHOR_PATTERNS = [re.compile(r"author")]
OTHER_PATTERNS = [re.compile(r"^other$"), re.compile(r"rolls on dmg magic item tables")]
GROUP_HEADING_PATTERNS = [
    re.compile(r"bundle", re.IGNORECASE),
    re.compile(r"side quests?:?$", re.IGNORECASE),
    re.compile(r"merchant princes", re.IGNORECASE),
    re.compile(r"items below for sale", re.IGNORECASE),
    re.compile(r"^guides$", re.IGNORECASE),
    re.compile(r"^lost tales of ", re.IGNORECASE),
]


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def normalize_text(value) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return f"{value.month}-{value.day}"

    if isinstance(value, date):
        return f"{value.month}-{value.day}"

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    return str(value).replace("\r", "").strip()


def normalize_code(value: str) -> str:
    return re.sub(r"\s+", " ", normalize_text(value)).strip()


def normalize_lookup(value: str) -> str:
    return normalize_header(normalize_text(value))


def compact_row_text(row):
    return " | ".join(value for value in (normalize_text(cell) for cell in row) if value)


def is_blank_row(row) -> bool:
    return not any(normalize_text(cell) for cell in row)


def score_header_row(row):
    score = 0
    normalized_values = [normalize_header(normalize_text(cell)) for cell in row if normalize_text(cell)]
    for value in normalized_values:
        for keyword in ADVENTURE_HEADER_KEYWORDS:
            if keyword in value:
                score += 1
    return score


def looks_like_subheader_row(row):
    values = [normalize_text(cell) for cell in row if normalize_text(cell)]
    if len(values) < 2 or len(values) > 12:
        return False

    if any(re.match(r"^(DD|CCC|SJ|WBW|FR|PS|SJA|DDEX)", value, re.IGNORECASE) for value in values):
        return False

    short_values = [value for value in values if len(value) <= 8]
    return len(short_values) / max(len(values), 1) >= 0.7


def dedupe_headers(headers):
    counts = {}
    deduped = []

    for index, header in enumerate(headers):
        base = header or f"Column {index + 1}"
        counts[base] = counts.get(base, 0) + 1
        count = counts[base]
        deduped.append(base if count == 1 else f"{base} {count}")

    return deduped


def build_headers(header_row, subheader_row):
    headers = []
    current_group = None
    current_group_count = 0

    max_length = max(len(header_row), len(subheader_row or []))

    for index in range(max_length):
        raw_header = normalize_text(header_row[index] if index < len(header_row) else "")
        raw_subheader = normalize_text(subheader_row[index] if subheader_row and index < len(subheader_row) else "")

        if raw_header:
            current_group = raw_header
            current_group_count = 1
            headers.append(f"{raw_header} {raw_subheader}".strip() if raw_subheader else raw_header)
            continue

        if current_group:
            current_group_count += 1
            headers.append(
                f"{current_group} {raw_subheader}".strip()
                if raw_subheader
                else f"{current_group} {current_group_count}"
            )
            continue

        headers.append(raw_subheader or f"Column {index + 1}")

    return dedupe_headers([re.sub(r"\s+", " ", header).strip() for header in headers])


def detect_header_info(values):
    search_rows = min(len(values), 6)
    best_index = 0
    best_score = -1

    for index in range(search_rows):
        score = score_header_row(values[index])
        if score > best_score:
            best_score = score
            best_index = index

    header_row = values[best_index] if values else []
    next_row = values[best_index + 1] if best_index + 1 < len(values) else None
    use_subheader = bool(next_row) and looks_like_subheader_row(next_row)

    return {
        "header_row_index": best_index,
        "subheader_row_index": best_index + 1 if use_subheader else None,
        "data_start_index": best_index + (2 if use_subheader else 1),
        "headers": build_headers(header_row, next_row if use_subheader else None),
    }


def map_row_to_object(headers, row):
    row_object = {}
    for index, header in enumerate(headers):
        row_object[header] = row[index] if index < len(row) else None
    return row_object


def first_match(row_object, patterns):
    for header, value in row_object.items():
        if not normalize_text(value):
            continue
        normalized = normalize_header(header)
        if any(pattern.search(normalized) for pattern in patterns):
            return value
    return None


def collect_matches(row_object, patterns):
    matches = []
    for header, value in row_object.items():
        if not normalize_text(value):
            continue
        normalized = normalize_header(header)
        if any(pattern.search(normalized) for pattern in patterns):
            matches.append((header, value))
    return matches


def looks_like_adventure_code(value) -> bool:
    text = normalize_code(value)
    if not text:
        return False
    return bool(re.match(r"^(DD|CCC|WBW|FR|PS|SJ|SJA|DDEX|DDAL|DRW|PO|LMOP|CM|GSM|DOIP|JRC|DSI|KGV)", text, re.IGNORECASE))


def excel_serial_to_month_day(serial_value):
    epoch = datetime(1899, 12, 30)
    as_date = epoch + timedelta(days=float(serial_value))
    return f"{as_date.month}-{as_date.day}"


def normalize_page_value(value):
    if isinstance(value, (datetime, date)):
        return f"{value.month}-{value.day}"

    if isinstance(value, (int, float)) and 20000 <= float(value) <= 60000:
        return excel_serial_to_month_day(value)

    return normalize_text(value)


def split_plain_list(value: str):
    text = normalize_text(value)
    if not text:
        return []

    parts = re.split(r"[\n;,]+", text.replace("\r\n", "\n"))
    return [part.strip() for part in parts if part.strip()]


def split_magic_items(value: str, known_items):
    text = normalize_text(value)
    if not text:
        return []

    results = []

    for chunk in re.split(r"[\n;]+", text):
        clean_chunk = chunk.strip()
        if not clean_chunk:
            continue

        normalized_chunk = normalize_lookup(clean_chunk)
        if normalized_chunk in known_items:
            results.append(clean_chunk)
            continue

        tokens = [part.strip() for part in clean_chunk.split(",") if part.strip()]
        if not tokens:
            continue

        index = 0
        while index < len(tokens):
            best_candidate = None
            best_size = 0
            max_window = min(4, len(tokens) - index)

            for size in range(max_window, 0, -1):
                candidate = ", ".join(tokens[index : index + size]).strip()
                if normalize_lookup(candidate) in known_items:
                    best_candidate = candidate
                    best_size = size
                    break

            if best_candidate:
                results.append(best_candidate)
                index += best_size
            else:
                results.append(tokens[index])
                index += 1

    return [item for item in results if item]


def append_unique(target_list, values):
    seen = {normalize_lookup(entry) for entry in target_list}
    for value in values:
        clean_value = normalize_text(value)
        if not clean_value:
            continue
        lookup = normalize_lookup(clean_value)
        if lookup in seen:
            continue
        target_list.append(clean_value)
        seen.add(lookup)


def sheet_family(sheet_name: str) -> str:
    if sheet_name in HELPER_SHEETS:
        return "helper"
    if sheet_name in ITEM_SHEETS:
        return "items"
    if re.match(r"^S\d+\s", sheet_name, re.IGNORECASE) or re.match(r"^Season ", sheet_name, re.IGNORECASE):
        return "seasonal_al"
    if re.search(r" HC$", sheet_name, re.IGNORECASE) or "Hardcover" in sheet_name:
        return "hardcover"
    if re.search(r" DCs$", sheet_name, re.IGNORECASE):
        return "dungeoncraft"
    if re.match(r"^CCC", sheet_name, re.IGNORECASE):
        return "ccc"
    if re.match(r"^Legacy$", sheet_name, re.IGNORECASE):
        return "legacy"
    if re.match(r"^(SJA|SJ LoX|SJ: LoX)$", sheet_name, re.IGNORECASE):
        return "spelljammer"
    return "adventure"


def tier_for_level(level_number):
    if 1 <= level_number <= 4:
        return "TIER_1"
    if 5 <= level_number <= 10:
        return "TIER_2"
    if 11 <= level_number <= 16:
        return "TIER_3"
    if 17 <= level_number <= 20:
        return "TIER_4"
    return ""


def parse_tiers(*values):
    tiers = []

    for raw_value in values:
        text = normalize_text(raw_value)
        normalized = normalize_header(text)

        if not normalized:
            continue

        if re.search(r"\btier 1\b|\bt1\b", normalized):
            append_unique(tiers, ["TIER_1"])
        if re.search(r"\btier 2\b|\bt2\b", normalized):
            append_unique(tiers, ["TIER_2"])
        if re.search(r"\btier 3\b|\bt3\b", normalized):
            append_unique(tiers, ["TIER_3"])
        if re.search(r"\btier 4\b|\bt4\b", normalized):
            append_unique(tiers, ["TIER_4"])

        for start_text, end_text in re.findall(r"(\d{1,2})\s*-\s*(\d{1,2})", normalized):
            start = int(start_text)
            end = int(end_text)

            if start <= 4 and end >= 1:
                append_unique(tiers, ["TIER_1"])
            if start <= 10 and end >= 5:
                append_unique(tiers, ["TIER_2"])
            if start <= 16 and end >= 11:
                append_unique(tiers, ["TIER_3"])
            if start <= 20 and end >= 17:
                append_unique(tiers, ["TIER_4"])

        numeric_values = [int(number) for number in re.findall(r"\b\d{1,2}\b", normalized)]
        if numeric_values and not re.findall(r"(\d{1,2})\s*-\s*(\d{1,2})", normalized):
            for number in numeric_values:
                mapped = tier_for_level(number)
                if mapped:
                    append_unique(tiers, [mapped])

    return [tier for tier in tiers if tier in {"TIER_1", "TIER_2", "TIER_3", "TIER_4"}]


def build_gold_summary(row_object):
    fixed = normalize_text(first_match(row_object, FIXED_GOLD_PATTERNS))
    rolled = normalize_text(first_match(row_object, ROLLED_GOLD_PATTERNS))
    legacy = normalize_text(first_match(row_object, LEGACY_GOLD_PATTERNS))

    populated = [value for value in [fixed, rolled, legacy] if value]
    if len(populated) == 1:
        return populated[0]

    parts = []
    if fixed:
        parts.append(f"Fixed: {fixed}")
    if rolled:
        parts.append(f"Rolled: {rolled}")
    if legacy:
        parts.append(f"Legacy: {legacy}")
    return " | ".join(parts)


def detect_heading_row(row_object) -> bool:
    values = [normalize_text(value) for value in row_object.values() if normalize_text(value)]
    if len(values) == 1:
        return True

    title_candidate = normalize_text(first_match(row_object, TITLE_PATTERNS))
    if title_candidate and any(pattern.search(title_candidate) for pattern in GROUP_HEADING_PATTERNS):
        return True

    return False


def build_rarity_map(workbook):
    rarity_map = {}

    automated_index = workbook["Automated Index"] if "Automated Index" in workbook.sheetnames else None
    if automated_index:
        rows = [[cell.value for cell in row] for row in automated_index.iter_rows()]
        for row in rows[1:]:
            if len(row) < 3:
                continue
            item_name = normalize_text(row[0])
            rarity = normalize_header(normalize_text(row[2]))
            if item_name and rarity:
                rarity_map[normalize_lookup(item_name)] = rarity

    unique_items_sheet = workbook["Unique Items"] if "Unique Items" in workbook.sheetnames else None
    if unique_items_sheet:
        rows = [[cell.value for cell in row] for row in unique_items_sheet.iter_rows()]
        if rows:
            headers = build_headers(rows[0], None)
            for row in rows[1:]:
                row_object = map_row_to_object(headers, row)
                display_name = normalize_text(row_object.get("Magic Item Name (if applicable)"))
                base_item = normalize_text(row_object.get("Base Item"))
                rarity = normalize_header(normalize_text(row_object.get("Rarity")))

                for candidate in [display_name, base_item]:
                    if candidate and rarity:
                        rarity_map[normalize_lookup(candidate)] = rarity

    return rarity_map


def classify_magic_items(magic_items, rarity_map):
    buckets = {
        "commonMagicItems": [],
        "uncommonMagicItems": [],
        "rareMagicItems": [],
        "veryRareMagicItems": [],
        "legendaryMagicItems": [],
        "uniqueMagicItems": [],
    }

    for item in magic_items:
        rarity = rarity_map.get(normalize_lookup(item), "")
        bucket_name = RARITY_BUCKETS.get(rarity, "uniqueMagicItems")
        append_unique(buckets[bucket_name], [item])

    return buckets


def empty_catalog_row(adventure_code, title, tier):
    return {
        "adventureCode": adventure_code,
        "title": title,
        "tier": tier,
        "duration": "",
        "goldParts": [],
        "spellbookParts": [],
        "storyAwardParts": [],
        "pageNumberParts": [],
        "sourceSheets": [],
        "sourceNotesParts": [],
        "consumables": [],
        "commonMagicItems": [],
        "uncommonMagicItems": [],
        "rareMagicItems": [],
        "veryRareMagicItems": [],
        "legendaryMagicItems": [],
        "uniqueMagicItems": [],
    }


def merge_catalog_rows(target, incoming):
    if incoming["duration"] and not target["duration"]:
        target["duration"] = incoming["duration"]

    append_unique(target["goldParts"], incoming["goldParts"])
    append_unique(target["spellbookParts"], incoming["spellbookParts"])
    append_unique(target["storyAwardParts"], incoming["storyAwardParts"])
    append_unique(target["pageNumberParts"], incoming["pageNumberParts"])
    append_unique(target["sourceSheets"], incoming["sourceSheets"])
    append_unique(target["sourceNotesParts"], incoming["sourceNotesParts"])
    append_unique(target["consumables"], incoming["consumables"])
    append_unique(target["commonMagicItems"], incoming["commonMagicItems"])
    append_unique(target["uncommonMagicItems"], incoming["uncommonMagicItems"])
    append_unique(target["rareMagicItems"], incoming["rareMagicItems"])
    append_unique(target["veryRareMagicItems"], incoming["veryRareMagicItems"])
    append_unique(target["legendaryMagicItems"], incoming["legendaryMagicItems"])
    append_unique(target["uniqueMagicItems"], incoming["uniqueMagicItems"])


def finalize_catalog_row(row):
    return {
        "adventureCode": row["adventureCode"],
        "title": row["title"],
        "tier": row["tier"],
        "duration": row["duration"],
        "gold": " | ".join(row["goldParts"]),
        "spellbook": " | ".join(row["spellbookParts"]),
        "storyAwards": " | ".join(row["storyAwardParts"]),
        "pageNumbers": " | ".join(row["pageNumberParts"]),
        "sourceSheet": ", ".join(row["sourceSheets"]),
        "sourceNotes": " | ".join(row["sourceNotesParts"]),
        "consumables": row["consumables"],
        "commonMagicItems": row["commonMagicItems"],
        "uncommonMagicItems": row["uncommonMagicItems"],
        "rareMagicItems": row["rareMagicItems"],
        "veryRareMagicItems": row["veryRareMagicItems"],
        "legendaryMagicItems": row["legendaryMagicItems"],
        "uniqueMagicItems": row["uniqueMagicItems"],
    }


def parse_catalog_from_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))

    if not rows:
        return []

    headers = [normalize_text(cell) for cell in rows[0]]
    columns = {normalize_header(header): index for index, header in enumerate(headers)}
    results = []

    def get_value(row, header_name):
        index = columns.get(normalize_header(header_name))
        if index is None or index >= len(row):
            return ""
        return normalize_text(row[index])

    for raw_row in rows[1:]:
        tier_values = parse_tiers(get_value(raw_row, "tier"))
        if not tier_values:
            continue

        results.append(
            {
                "adventureCode": get_value(raw_row, "adventure code"),
                "title": get_value(raw_row, "adventure title"),
                "tier": tier_values[0],
                "duration": get_value(raw_row, "duration"),
                "gold": get_value(raw_row, "gold"),
                "spellbook": get_value(raw_row, "spellbook"),
                "storyAwards": get_value(raw_row, "story awards"),
                "pageNumbers": "",
                "sourceSheet": path.stem,
                "sourceNotes": "Imported from CSV source.",
                "consumables": split_plain_list(get_value(raw_row, "consumables")),
                "commonMagicItems": split_plain_list(get_value(raw_row, "common magic items")),
                "uncommonMagicItems": split_plain_list(get_value(raw_row, "uncommon magic items")),
                "rareMagicItems": split_plain_list(get_value(raw_row, "rare magic items")),
                "veryRareMagicItems": split_plain_list(get_value(raw_row, "very rare magic items")),
                "legendaryMagicItems": split_plain_list(get_value(raw_row, "legendary magic items")),
                "uniqueMagicItems": split_plain_list(get_value(raw_row, "unique magic items")),
            }
        )

    return results


def parse_catalog_from_workbook(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    rarity_map = build_rarity_map(workbook)
    known_items = set(rarity_map.keys())
    aggregates = OrderedDict()

    for sheet_name in workbook.sheetnames:
        if sheet_name in HELPER_SHEETS or sheet_name in ITEM_SHEETS:
            continue

        worksheet = workbook[sheet_name]
        values = [[cell.value for cell in row] for row in worksheet.iter_rows()]
        if not values:
            continue

        header_info = detect_header_info(values)
        headers = header_info["headers"]
        data_start_index = header_info["data_start_index"]

        current_code = ""
        current_title = ""
        current_tiers = []
        current_duration = ""

        for row_index in range(data_start_index, len(values)):
            row = values[row_index]
            if is_blank_row(row):
                continue

            row_object = map_row_to_object(headers, row)
            row_code_candidate = first_match(row_object, CODE_PATTERNS)
            row_title_candidate = first_match(row_object, TITLE_PATTERNS)
            chapter_title_candidate = first_match(row_object, [re.compile(r"chapter section title"), re.compile(r"chapter title")])

            row_code = normalize_code(row_code_candidate) if looks_like_adventure_code(row_code_candidate) else ""
            row_title = normalize_text(row_title_candidate) or normalize_text(chapter_title_candidate)
            level_range = normalize_text(first_match(row_object, LEVEL_PATTERNS))
            recommended_level = normalize_text(first_match(row_object, RECOMMENDED_PATTERNS))
            apl_value = normalize_text(first_match(row_object, APL_PATTERNS))
            duration_value = normalize_text(first_match(row_object, RUNTIME_PATTERNS))
            page_numbers = normalize_page_value(first_match(row_object, PAGE_NUMBER_PATTERNS))
            story_award_values = [normalize_text(value) for _, value in collect_matches(row_object, STORY_AWARD_PATTERNS)]
            other_values = [f"{header}: {normalize_text(value)}" for header, value in collect_matches(row_object, OTHER_PATTERNS)]
            spellbook_values = [normalize_text(value) for _, value in collect_matches(row_object, SPELLBOOK_PATTERNS)]
            consumable_values = []
            for _, value in collect_matches(row_object, CONSUMABLE_PATTERNS):
                consumable_values.extend(split_plain_list(value))

            magic_items = []
            for _, value in collect_matches(row_object, MAGIC_ITEM_PATTERNS):
                magic_items.extend(split_magic_items(value, known_items))

            if row_code:
                current_code = row_code
            if row_title:
                current_title = row_title

            parsed_tiers = parse_tiers(level_range, recommended_level, apl_value, row_code)
            if parsed_tiers:
                current_tiers = parsed_tiers

            if duration_value:
                current_duration = duration_value

            effective_code = row_code or current_code
            effective_title = row_title or current_title
            effective_tiers = parsed_tiers or current_tiers
            effective_duration = duration_value or current_duration

            if detect_heading_row(row_object) and not effective_code:
                continue

            has_row_content = any(
                [
                    effective_duration,
                    page_numbers,
                    build_gold_summary(row_object),
                    story_award_values,
                    spellbook_values,
                    consumable_values,
                    magic_items,
                    row_code,
                    row_title,
                ]
            )

            if not effective_code or not effective_title or not effective_tiers or not has_row_content:
                continue

            classified_magic_items = classify_magic_items(magic_items, rarity_map)
            source_notes = []
            author = normalize_text(first_match(row_object, AUTHOR_PATTERNS))
            gold_summary = build_gold_summary(row_object)

            if author:
                source_notes.append(f"Author: {author}")
            append_unique(source_notes, other_values)
            if not row_code and current_code:
                source_notes.append("Adventure code inherited from the previous related row.")
            if not row_title and current_title:
                source_notes.append("Adventure title inherited from the previous related row.")

            for tier in effective_tiers:
                lookup_key = (normalize_lookup(effective_code), normalize_lookup(effective_title), tier)

                if lookup_key not in aggregates:
                    aggregates[lookup_key] = empty_catalog_row(effective_code, effective_title, tier)

                incoming = empty_catalog_row(effective_code, effective_title, tier)
                incoming["duration"] = effective_duration
                append_unique(incoming["goldParts"], [gold_summary])
                append_unique(incoming["spellbookParts"], spellbook_values)
                append_unique(incoming["storyAwardParts"], story_award_values)
                append_unique(incoming["pageNumberParts"], [page_numbers])
                append_unique(incoming["sourceSheets"], [sheet_name])
                append_unique(incoming["sourceNotesParts"], source_notes)
                append_unique(incoming["consumables"], consumable_values)

                for bucket_name, items in classified_magic_items.items():
                    append_unique(incoming[bucket_name], items)

                merge_catalog_rows(aggregates[lookup_key], incoming)

    return [finalize_catalog_row(row) for row in aggregates.values()]


def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: extract-adventures-from-workbook.py <spreadsheet-path>")

    file_path = Path(sys.argv[1]).expanduser().resolve()

    if not file_path.exists():
        raise SystemExit(f"Workbook not found: {file_path}")

    if file_path.suffix.lower() == ".csv":
        parsed_rows = parse_catalog_from_csv(file_path)
    else:
        parsed_rows = parse_catalog_from_workbook(file_path)

    print(json.dumps(parsed_rows))


if __name__ == "__main__":
    main()
